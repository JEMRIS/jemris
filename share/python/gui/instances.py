"""
This program is free software; you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation; either version 2 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program; if not, write to the Free Software
Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA  02110-1301  USA
"""

import os
import copy
import string
import numpy as np

import h5py
import nibabel as nib
import openpyxl

import xml.etree.ElementTree as ET

from gui.settings import workspace
from gui.logs import logger
from gui.parser import parseXML, writeXML


# Sequence Element
class Element(object):

    def __init__(self, attr, xml=None):

        self.attr = attr

        if xml is not None:
            self.xml = xml
        else:
            self.xml = self.newItem('PARAMETERS', attr=self.attr)

        # Classify all entries.
        for element in [self.xml] + self.xml.findall('.//'):
            element.set('type', classifyEntry(element.tag))
            element.set('color', colorEntry(element.get('type')))

    def get(self):

        return self.xml

    def __access(self, idx, mode=None, item=None):
        
        ref = [[self.xml]]          # double braces to include tree root
        for i in idx[:-1]:          # single braces for standard access
            ref.append(ref[-1][i])
        node = ref[-1]

        if mode == 'read':
            return node[idx[-1]]
        elif mode == 'delete':
            del(node[idx[-1]])
        elif mode == 'write':
            node[idx[-1]] = item
        elif mode == 'insert':
            node.insert(idx[-1], item)
        elif mode == 'append':
            node[idx[-1]].append(item)

    def read(self, idx):

        return self.__access(idx, mode='read', item=None)

    def write(self, idx, item):

        self.__access(idx, mode='write', item=item)

    def delete(self, idx):

        self.__access(idx, mode='delete', item=None)

    def append(self, idx, item):

        self.__access(idx, mode='append', item=item)

    def insert(self, idx, item):

        self.__access(idx, mode='insert', item=item)

    def copy(self, idx):

        item = copy.deepcopy(self.read(idx))
        self.insert(idx, item)

    def _move(self, dragIndex, dropIndex):

        item = self.read(dragIndex)
        self.delete(dragIndex)
        self.insert(dropIndex, item)

    def move(self, dragIndex, dropIndex):

        # SAVE INSERT #
        tmp = copy.deepcopy(self)
        tmp._move(dragIndex, dropIndex)
        status, msg = tmp.checkCommand(dropIndex)
        if status:
            self._move(dragIndex, dropIndex)
        else:
            return msg

    def _insert(self, idx, tag):

#        item = self.newItem(tag, self.attr, self.xml)
        item = self.newItem(tag, self.attr)  #dridder

        # Automatic Naming
        item.set('type', classifyEntry(item.tag))
        item.set('color', colorEntry(item.get('type')))

        number = 1
        for element in [self.xml] + self.xml.findall('.//'):
            if element.get('type') == item.get('type'):
                number += 1
        item.attrib['Name'] = '%s%i' % (item.get('type')[0], number)

        # SAVE INSERT #
        tmp = copy.deepcopy(self)
        tmp.append(idx, item)
        status, msg = tmp.checkCommand(idx + [-1])
        if status:
            self.append(idx, item)
        else:
            return msg

    def newItem(self, tag, attr):

        " Create new Element "
        item = ET.Element(tag)
        match = attr.find('.//' + tag.upper())

        for key in match.attrib.keys():
            if match.attrib[key] not in ['', '0']:
                item.attrib[key] = match.attrib[key]

        item.attrib['Name'] = item.tag[0]   # First Letter

        return item

    def returnAttr(self, idx, showHidden):

        attr = self.attr

        item = self.read(idx)
        match = attr.find('.//' + item.tag.upper())

        if match is None:
            return [], [], item.tag.upper()

        # Sort keys, Name and Observe must be first
        keys = sorted(match.attrib.keys())

        for idx, key in enumerate(['Name', 'Observe']):
            if key in keys:
                keys.remove(key)
                keys.insert(idx, key)

        if showHidden:
            keys = [key.rstrip('HIDDEN') for key in keys]
        else:
            keys = [key for key in keys if key.find('HIDDEN') is -1]

        values = list()
        for key in keys:
            if key in item.attrib.keys():
                values.append(item.attrib[key])
            else:
                values.append(None)

        return keys, values, match.tag

    def writeAttr(self, idx, key, value):

        item = self.read(idx)
        item.attrib[key] = value

        return item.attrib['Name'], item.tag

    def checkCommand(self, idx):

        type_child = self.read(idx).get('type')

        " Parameter node cannot be move "
        if type_child in ['PARAMETERS', 'COILARRAY']:
            return [False, 'Head node cannot be moved.']

#        if len(idx):
#            return [False, 'Head node cannot be moved.']

        flag = False
        msg = None
        type_parent = self.read(idx[:-1]).get('type')

        pairs = [{'child':  ['PULSE'],
                  'parent': ['ATOMICSEQUENCE']},
                 {'child':  ['CONCATSEQUENCE',
                             'ATOMICSEQUENCE',
                             'DELAYATOMICSEQUENCE'],
                  'parent': ['PARAMETERS',
                             'CONCATSEQUENCE']},
                 {'child':  ['COIL'],
                  'parent': ['COILARRAY']}]
        for pair in pairs:
            if type_child in pair['child']:
                if type_parent in pair['parent']:
                    flag = True
                else:
                    msg = ('Element of Type %s can '
                           'only be inserted below Elements of type %s') \
                          % (type_child, ', '.join(pair['parent']))

        return flag, msg


def classifyEntry(tag):

    tag = tag.upper()

    if tag.find('PULSE') is not -1:

        _type = 'PULSE'

    elif tag in ['BIOTSAVARTLOOP', 'ANALYTICCOIL', 'EXTERNALCOIL']:

        _type = 'COIL'

    else:

        _type = tag

    return _type


def colorEntry(_type):

    _color = {'CONCATSEQUENCE':         'yellow',
              'ATOMICSEQUENCE':         'blue',
              'DELAYATOMICSEQUENCE':    'green',
              'PULSE':                  'white',
              'PARAMETERS':             'red',
              'COILARRAY':              'red',
              'COIL':                   'white'}

    return _color[_type]


# TODO: Merge Sequence and Array class.
class Sequence(object):

    def __init__(self, attr, fname=None):

        self.attr = attr
        self.data = None

        if fname is not None:
            self.rename(fname)
        else:
            self.rename(os.path.join(workspace(), 'new_sequence.xml'))
            self.tmpname = True
            self.xml = Element(self.attr, None)

    def rename(self, fname):

        self.tmpname = False
        # strip extension
        self.fname = os.path.splitext(fname)[0]
        self.PathName, self.FileName = os.path.split(self.fname)

    def readXML(self):
        
        print(self.fname)
        
        self.xml = Element(self.attr, parseXML('%s.xml' % self.fname))
        # convert all tags to upper case to avoid classification problems
        for element in [self.xml.get()] + self.xml.get().findall('.//'):
            element.tag = element.tag.upper()

    def save(self):

        # current inconsistency in JEMRIS
        for element in [self.xml.get()] + self.xml.get().findall('.//'):
            if element.tag == 'PARAMETERS':
                element.tag = 'Parameters'
            if element.tag == 'CONCATSEQUENCE':
                element.tag = 'ConcatSequence'
        # Throw away hidden attributes
        # TODO: Should they be editable at all?
            for key in element.keys():
                if key.find('HIDDEN') is not -1:
                    del (element.attrib[key])

        writeXML('%s.xml' % self.fname, self.xml.get())

    def make(self, caller, fcn):

        if self.fname is None:
            return

        def finished():
            self.readh5()
            if fcn is not None:
                fcn()

        caller.make(self, finished)

    def readh5(self):

        if os.path.exists('%s.h5' % self.fname):

            """
            T:          Time Points
            RXP:        Receive phase
            TXM, TXP:   Transmit magnitude and phase
            GX, GY, GZ: Gradients
            KX, KY, KZ: k-Space Points
            META:       Unknown
            """

            file = h5py.File('%s.h5' % self.fname, 'r')

            A = dict()

            for key in file['/seqdiag'].keys():
                A[key] = np.array(file['/seqdiag/%s' % key])

            A['J']       = np.where(A['META']>2)[0]     # k-lines
            A['Iadc']    = np.where(A['RXP']>=0)[0]     # ADC index
            A['Tadc']    = A['T'][A['Iadc']]            # ADC time points
            A['Rec_Phs'] = A['RXP'][A['Iadc']]          # ADC phase locks

            # Radians to Degrees [-180, 180]
            for key in ['RXP', 'TXP', 'Rec_Phs']:
                A[key] = np.rad2deg(np.angle(np.exp(1j*A[key])))

            self.data = A

            self.status = 0

        else:

            logger.info('%s.h5 does not exist.' % self.fname)

            self.status = 1


class Array(object):

    def __init__(self, attr, fname=None):

        self.attr = attr
        self.data = None

        if fname is not None:
            self.rename(fname)
        else:
            self.rename(os.path.join(workspace(), 'new_array.xml'))
            self.tmpname = True

    def rename(self, fname):

        self.tmpname = False
        # strip extension
        self.fname = fname = os.path.splitext(fname)[0]
        self.PathName, self.FileName = os.path.split(self.fname)

    def save(self):

        writeXML('%s.xml' % self.fname, self.xml.get())

    def readXML(self):

        self.xml = Element(self.attr, parseXML('%s.xml' % self.fname))

    def make(self, caller, fcn):

        if self.fname is None:
            return

        def finished():
            self.readh5()
            if fcn is not None:
                fcn()

        caller.make(self, finished)

    def readh5(self):

        if os.path.exists('%s.h5' % self.fname):

            file = h5py.File('%s.h5' % self.fname, 'r')

            A = list()

            for key in file['/maps/magnitude'].keys():
                m = np.array(file['/maps/magnitude/%s' % key])
                p = np.array(file['/maps/phase/%s' % key])

                A.append(m * np.exp(1j*p))

            A = np.array(A)

            if len(np.shape(A)) < 4:
                A = A[:, None, :, :]

            A = A.transpose([3, 2, 1, 0])

            A = np.append(A, np.sum(A, axis=-1)[..., None], axis=-1)

            self.data = A

        else:

            logger.info('%s.h5 does not exist.' % self.fname)

            self.status = 1


class Simulation(object):

    def __init__(self, fname):

        # Defaults.
        model = {'name':    'Bloch',
                 'type':    'CVODE'}

        parameter = {'ConcomitantFields':   '0',
                     'EvolutionPrefix':     'evol',
                     'EvolutionSteps':      '0',
                     'RandomNoise':         '0'}

#        self.sim = dict()
        self.sim = {'sample':       dict(),
                    'sequence':     dict(),
                    'RXcoilarray':  dict(),
                    'TXcoilarray':  dict(),
                    'model':        model,
                    'parameter':    parameter}

        self.data = None
        self.channels = []

        if fname is not None:
            self.rename(fname)
        else:
            self.rename(os.path.join(workspace(), 'new_simulation.xml'))
            self.tmpname = True
            self.setSequence('')
            self.setSample('')
            self.setTxCoilArray('')
            self.setRxCoilArray('')

    def rename(self, fname):

        self.tmpname = False
        # strip extension
        self.fname = os.path.splitext(fname)[0]
        self.PathName, self.FileName = os.path.split(self.fname)

    def setSequence(self, sequenceURI):

        self.sim['sequence']['uri'] = sequenceURI
        self.sim['sequence']['name'] = os.path.split(sequenceURI)[1]
        return self.sim['sequence']['name']

    def getSequence(self):

        return self.sim['sequence']['uri']

    def setSample(self, sampleURI):

        self.sim['sample']['uri'] = sampleURI
        self.sim['sample']['name'] = os.path.split(sampleURI)[1]
        return self.sim['sample']['name']

    def getSample(self):

        return self.sim['sample']['uri']

    def setR2Trajectory(self, uri):

        self.sim['sample']['R2Trajectory'] = uri

    def getR2Trajectory(self):

        return self.sim['sample']['R2Trajectory']

    def setMotionTrajectory(self, uri):

        self.sim['sample']['MotionTrajectory'] = uri

    def getMotionTrajectory(self):

        return self.sim['sample']['MotionTrajectory']

    def setDiffusionFile(self, value):

        self.sim['sample']['Diffusionfile'] = value

    def getDiffusionFile(self):

        return self.sim['sample']['Diffusionfile']

    def setActiveCircles(self, value):

        self.sim['sample']['ActiveCircles'] = value

    def getActiveCircles(self):

        return self.sim['sample']['ActiveCircles']

    def setTxCoilArray(self, txcoilarrayURI):

        self.sim['TXcoilarray']['uri'] = txcoilarrayURI
        self.sim['TXcoilarray']['name'] = os.path.split(txcoilarrayURI)[1]
        return self.sim['TXcoilarray']['name']

    def getTxCoilArray(self):

        return self.sim['TXcoilarray']['uri']

    def setRxCoilArray(self, rxcoilarrayURI):

        self.sim['RXcoilarray']['uri'] = rxcoilarrayURI
        self.sim['RXcoilarray']['name'] = os.path.split(rxcoilarrayURI)[1]
        return self.sim['RXcoilarray']['name']

    def getRxCoilArray(self):

        return self.sim['RXcoilarray']['uri']

    def readXML(self):

        root = parseXML('%s.xml' % self.fname)
        for child in root.getchildren():
            for key in child.attrib.keys():
                self.sim[child.tag][key] = child.attrib[key]

    def save(self):

        root = ET.Element('simulate', attrib={'name': 'JEMRIS'})

        for key, value in self.sim.iteritems():
            root.append(ET.Element(tag=key, attrib=value))

        writeXML('%s.xml' % self.fname, root)

    def make(self, caller, mode, fcn):

        if self.fname is None:
            return

        def finished():
            self.readh5()
            if fcn is not None:
                fcn()

        caller.make(self, finished, mode)

    def readh5(self):

        if os.path.exists('%s.h5' % self.fname):

            file = h5py.File('%s.h5' % self.fname, 'r')

            tmp = '/signal/channels'

            self.channels = file[tmp].keys()

            self.data = dict()
            self.data['T'] = np.array(file['/signal/times'])
            self.data['S'] = list()

            for key in self.channels:
                self.data['S'].append(np.array(file['%s/%s' % (tmp, key)]))

            # Time, Orientation (Mx, My, Mz), Channel
            self.data['S'] = np.array(self.data['S']).transpose([1, 2, 0])

            self.status = 0

        else:

            logger.info('%s.h5 does not exist.' % self.fname)

            self.status = 1


class Phantom(object):

    def __init__(self, fname):

        # = DxDy (sometimes given as scalar)
        self.RES = np.array([1., 1., 1.])[:, None]
        # is a shift in each dimension
        self.OFF = np.array([0., 0., 0.])[:, None]

        self.labelPhantom = dict()
        self.labelPhantom['active'] = False
        self.labelMatrix = dict()
        self.labelMatrix['active'] = False
        self.mapsComputed = False

        self.sfacs = dict()     # scaling factors
        for key in ['M0', 'T1', 'T2', 'T2s', 'CS', 'Xi']:
            self.sfacs[key] = 1

        if fname is not None:
            self.rename(fname)
        else:
            self.rename(os.path.join(workspace(), 'new_phantom.h5'))
            self.tmpname = True

        self.geometricPhantoms = ['Sphere', 'Spheres', 'Cylinder', 'Revolver']

    def rename(self, fname):

        self.tmpname = False
        # strip extension
        self.fname = fname = os.path.splitext(fname)[0]
        self.PathName, self.FileName = os.path.split(self.fname)

    def setRange(self, _range=None):

        if _range is None:
            self.range = [[0, self.labelPhantom['shape'][0]],
                          [0, self.labelPhantom['shape'][1]],
                          [0, self.labelPhantom['shape'][2]]]
        else:
            self.range = [[r-1 for r in ra] for ra in _range]
    #        self.sliceRange = np.arange(sliceRange[0]-1, sliceRange[1])
            print(self.range)

    def setLabelPhantom(self, _str, isfile):

        if isfile:
            nii = nib.load(_str)
            mask = nii.get_data()
        else:
            if _str == 'Sphere':     # homogenous sphere
                mask = sphericalPhantom(R=[0, 50], L=[1])
            elif _str == 'Spheres':
                mask = sphericalPhantom(R=[0, 25, 50], L=[1, 2])
            elif _str == 'Cylinder':
                mask = cylindricalPhantom(radius=50, length=100)
            elif _str == 'Revolver':
                mask = revolverPhantom()

        self.labelPhantom['data'] = mask
        self.labelPhantom['shape'] = np.shape(mask)
        self.labelPhantom['active'] = True
        self.labelPhantom['set'] = set(np.unique(mask))

        self.setRange()
        self.mapsComputed = False

    def setLabelMatrix(self, fname):

        self.wbname = fname

        " Read xlsx File and create Dictionary "
        workbook = openpyxl.load_workbook(self.wbname, use_iterators=True)
        sheet = workbook.get_active_sheet()
#        sheet = workbook.get_sheet_by_name(name=self.shname)
        _dict = dict()

        try:
            # Old version
            max_col = string.ascii_uppercase.find(sheet.max_col)
        except:
            # New version
            max_col = sheet.max_column-1

        # NOTE: Inconsistency between platforms
        # row[index].value and row[index].internal_value
        # Fix by switching modules or even from xlsx to odf?
        for row in sheet.iter_rows():
            try:
                _list = [row[i].value for i in range(1, max_col+1)]
                _dict[str(row[0].value)] = _list
            except:
                _list = [row[i].internal_value for i in range(1, max_col+1)]
                _dict[str(row[0].internal_value)] = _list

        " Create Matrix From Dictionary "
        header = ['None', '_LABEL', '_UNIT']
        tissue = [_dict[key] for key in _dict.keys() if key not in header]
        materials = [key for key in _dict.keys() if key not in header]
        tissue = np.array(tissue)                   # convert to numpy array
        materials = [x for (y, x) in sorted(zip(tissue[:, -1], materials))]
        tissue = tissue[tissue[:, -1].argsort()]    # sort by labels column

        self.labelMatrix['materials'] = materials
        self.labelMatrix['matrix'] = tissue
        self.labelMatrix['set'] = set(np.unique(self.labelMatrix['matrix'][:, -1]))
        self.labelMatrix['active'] = True
        self.labelMatrix['labels'] = [str(_str) for _str in _dict['_LABEL']]
        self.labelMatrix['labels'][-1] = 'Label'
        self.labelMatrix['units'] = [str(_str) for _str in _dict['_UNIT']]

        self.mapsComputed = False

    def computeParameterMaps(self):

        if not(self.labelPhantom['active'] and self.labelMatrix['active']):
            return

        mask = self.labelPhantom['data']
        TM = self.labelMatrix['matrix']
        labels_in_TM = self.labelMatrix['set']
        labels_in_mask = self.labelPhantom['set']

        if not labels_in_mask.issubset(labels_in_TM):
            raise ValueError('Mask contains unspecified labels.')

        maps = np.zeros((np.size(mask), np.shape(TM)[1] - 1))

        for label in labels_in_mask:
            values = TM[TM[:, -1] == label, :-1]
            maps[mask.ravel() == label, :] = values

        self.maps = dict()      # raw maps
        for idx, key in enumerate(['M0', 'T1', 'T2', 'T2s', 'CS', 'Xi']):
            self.maps[key] = maps[:, idx].reshape(np.shape(mask))

        self.maps['CS'] *= 2*np.pi
        self.XB = computeXB(self.maps['Xi'])

        self.mapsComputed = True

    def computeScaledMaps(self):

        self.smaps = dict()     # scaled maps

        for key in self.maps.keys():
            self.smaps[key] = copy.deepcopy(self.maps[key])[self.range[0][0]:self.range[0][1],
                                                            self.range[1][0]:self.range[1][1],
                                                            self.range[2][0]:self.range[2][1]] * self.sfacs[key]

        # TODO: Is DB really DB or actually DF (unit is MHz)
        self.DB = self.smaps['CS'] + self.sfacs['Xi'] * self.XB[self.range[0][0]:self.range[0][1],
                                                                self.range[1][0]:self.range[1][1],
                                                                self.range[2][0]:self.range[2][1]]

    def save(self):

        # NN = np.zeros(np.shape(M0))

        DATA = [self.smaps['M0'],
                self.smaps['T1'],
                self.smaps['T2'],
                self.smaps['T2s'],
                self.DB]

        DATA = np.array(DATA)
        with np.errstate(divide='ignore'):
            for i in range(1, 4):
                DATA[i, ...] = 1./DATA[i, ...]  # Invert Time Constants
        DATA[~np.isfinite(DATA)] = 0

        " NUMPY [p, x, y, z] <-> JEMRIS [z, y, x, p] "
        DATA = DATA.transpose([3, 2, 1, 0])

        name = ['/sample/%s' % n for n in ['data',
                                           'resolution',
                                           'offset']]
        data = [DATA, self.RES, self.OFF]

        with h5py.File('%s.h5' % self.fname, 'w') as f:
            for n, d in zip(name, data):
                f.create_dataset(n, d.shape, d.dtype)[:] = d

    def readh5(self):

        with h5py.File('%s.h5' % self.fname, 'r') as f:

            DATA = np.array(f["/sample/data"])
            self.RES = np.array(f["/sample/resolution"])
            self.OFF = np.array(f["/sample/offset"])

        DATA = DATA.transpose([3, 2, 1, 0])
        with np.errstate(divide='ignore'):
            for i in range(1, 4):
                DATA[i, ...] = 1./DATA[i, ...]  # Invert Time Constants
        DATA[~np.isfinite(DATA)] = 0

        self.maps = dict()
        self.maps['M0'] = DATA[0, ...]
        self.maps['T1'] = DATA[1, ...]
        self.maps['T2'] = DATA[2, ...]
        self.maps['T2s'] = DATA[3, ...]

        self.maps['CS'] = np.zeros(self.maps['M0'].shape)
        self.maps['Xi'] = np.zeros(self.maps['M0'].shape)
        self.XB = DATA[4, ...]

        self.mapsComputed = True

        self.labelPhantom['shape'] = np.shape(self.maps['M0'])
        self.setRange()


def computeXB(Xi):

    """
    Formerly "CalcBfromXi"
    S: Susceptibility Map
    """

    B0 = 1.5                            # Static magnetic field [T]
    gamma = 42.576 * B0 * 1e6           # Gyromagnetic Ratio    [MHz/T]

    print('Computing XB')

    N = np.shape(Xi)
    M = (2**np.ceil(np.log2(N))).astype(int)
#        M = 2*M # for increased precision (takes long)
    _Xi = np.zeros(M)
    _Xi[0:N[0], 0:N[1], 0:N[2]] = Xi
    Xi = _Xi * 1e-6     # given in ppm

    # differential operator in frequency domain
    [Nx, Ny, Nz] = np.shape(Xi)

    fx, fy, fz = [np.linspace(0, 0.5, n/2+1) for n in [Nx, Ny, Nz]]
    fx, fy, fz = [np.concatenate((f, -f[-2:0:-1])) for f in [fx, fy, fz]]

    K = np.array(np.meshgrid(fx, fy, fz))   # [Kx, Ky, Kz]

    Dk = 1./3 - (K[-1, ...]**2 / np.sum(K**2, axis=0))
    Dk[0, 0, 0] = 1./3

    B = np.real(np.fft.ifftn(Dk*np.fft.fftn(Xi, norm='ortho'), norm='ortho'))

    B = B[0:N[0], 0:N[1], 0:N[2]]       # Magnetic field distortion [T]

    print('Finished XB')

    return B*B0*2*np.pi*gamma


def sphericalPhantom(R, L):

    N = 2*max(R)
    mask = np.zeros((N, N, N))

    x, y, z = np.mgrid[-N/2+1:N/2, -N/2+1:N/2, -N/2+1:N/2]

    norm = np.sqrt(x**2 + y**2 + z**2)

    for idx, l in enumerate(L):
        mask[np.where((R[idx] <= norm) & (norm < R[idx+1]))] = l

    return mask


def cylindricalPhantom(radius, length):

    X, Y, Z = [2*radius, 2*radius, length]

    phantom = np.zeros((X, Y, Z))

    x, y, z = np.mgrid[-X/2+1:X/2, -Y/2+1:Y/2, -Z/2+1:Z/2]

    norm = np.sqrt(x**2 + y**2)

    phantom[np.where(norm < radius )] = 1

    return phantom


def revolverPhantom():

#    import numpy as np
#    import matplotlib.pyplot as plt
    X, Y, Z = [100, 100, 100]

    phantom = np.zeros((X, Y, Z))

    wt = 3                      # Wall Thickness
    sp = 5                      # Air Spacing around Probes
    Rout = X/2 - 1              # Outer Ring
    Rin = 1./3 * (2*Rout-wt)    # Inner Ring
    r2 = 1./3 * (Rout-2*wt)     # Probe Spacing
    r1 = r2 - sp                # Probe Radius

    x, y, z = np.mgrid[-X/2+1:X/2, -Y/2+1:Y/2, -Z/2+1:Z/2]

    phi = np.radians(np.array((30, 90, 150, 210, 270, 330)))
    posx = Rin*np.cos(phi)
    posy = Rin*np.sin(phi)

    label = [1, 2, 3, 4, 5, 6]

    norm = np.sqrt((x)**2 + (y)**2)
    phantom[np.where(norm < Rout)] = 7

    for px, py, l in zip(posx, posy, label):
        norm = np.sqrt((x-px)**2 + (y-py)**2)
        phantom[np.where(norm < r1)] = l
        phantom[np.where((norm > r1) & (norm < r2))] = 0

#    plt.imshow(phantom[:,:,0])

    return phantom