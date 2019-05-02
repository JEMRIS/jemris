# -*- coding: utf-8 -*-
"""
Created on Thu Sep  3 15:53:14 2015

@author: user
"""


import numpy as np
import openpyxl
import nibabel as nib

fname_label = '/home/user/Desktop/MNIbrain_corr.nii'
fname_phantom = 'phantom.nii'


def tissueMatrix():

    fname = 'tissue_classes.xlsx'
    sname = 'healthy'

    " Read xlsx File and create Dictionary "
    wb = openpyxl.load_workbook(fname, use_iterators=True)
    sh = wb.get_sheet_by_name(name=sname)
    dt = dict()
    for row in sh.iter_rows():
        dt[str(row[0].value)] = [row[i].value for i in range(1, sh.max_column)]
        # On MAC use internal_value instead of value
        # Also max_column does not work

    " Create Matrix From Dictionary "
    tissue = [dt[key] for key in dt.keys() if key != 'None']
    tissue = np.array(tissue)                   # convert to numpy array
    tissue = tissue[tissue[:, -1].argsort()]    # sort by labels column
    tissue = tissue[:, :-1]                     # remove labels column

    return tissue

nii = nib.load(fname_label)

mask = nii.get_data()

tM = tissueMatrix()
phantom = np.zeros((np.size(mask), np.shape(tM)[1]))

for idx, voxel in enumerate(mask.ravel()):

    phantom[idx, :] = tM[voxel, :]

phantom = phantom.reshape(np.shape(mask) + (np.shape(tM)[1],))

nii = nib.Nifti1Image(phantom, np.eye(4))
nib.save(nii, fname_phantom)
